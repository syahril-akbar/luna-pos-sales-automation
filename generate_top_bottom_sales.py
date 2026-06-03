"""
Generate Top 20 Penjualan Produk Tertinggi
Source: Ringkasan Penjualan Produk 01 April - 30 April 2026.xlsx
Output:
  - Top_20_Penjualan_Tertinggi_April_2026.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os
import sys
import argparse
import re

# ─────────────────────────────────────────────
# Daftar kategori yang diizinkan (urutan = prioritas matching)
# Setiap entry: (label_kategori, fungsi_pengecekan_nama_produk)
# ─────────────────────────────────────────────
CATEGORY_RULES = [
    ("Bubur 9+ 100ml",          lambda n: "BUBUR" in n and "9+"  in n and "100" in n and "ML" in n and "MEAL BOX" not in n),
    ("Bubur 6+ 200ml",          lambda n: "BUBUR" in n and "6+"  in n and "200" in n and "ML" in n and "MEAL BOX" not in n),
    ("Sup",                     lambda n: "SUP " in n or n.startswith("SUP")),
    ("Bubur 6+ 100ml",          lambda n: "BUBUR" in n and "6+"  in n and "100" in n and "ML" in n and "MEAL BOX" not in n),
    ("Bubur 9+ 200ml",          lambda n: "BUBUR" in n and "9+"  in n and "200" in n and "ML" in n and "MEAL BOX" not in n),
    ("Bubur 11+ 100ml",         lambda n: "BUBUR" in n and "11+" in n and "100" in n and "ML" in n and "MEAL BOX" not in n),
    ("Bubur 11+ 200ml",         lambda n: "BUBUR" in n and "11+" in n and "200" in n and "ML" in n and "MEAL BOX" not in n),
    ("Lauk",                    lambda n: "LAUK" in n),
    ("Snack Buah",              lambda n: "SNACK BUAH" in n),
    ("Rice BB Booster",         lambda n: ("BUTTER RICE" in n or "SEAWEED BUTTER RICE" in n
                                           or ("RICE" in n and "BUBUR" not in n
                                               and "RICE BOX" not in n
                                               and "YAKINIKU" not in n))),
    ("Finger Food",             lambda n: "FINGER FOOD" in n),
    ("Kaldu BB Booster",        lambda n: "KALDU" in n),
    ("Abon 25ml",               lambda n: "ABON" in n and "25ML" in n),
    ("Abon 10ml",               lambda n: "ABON" in n and "25ML" not in n),
    ("Rice Box",                lambda n: "RICE BOX" in n),
    ("Pasta",                   lambda n: "PASTA" in n),
    ("Kremes",                  lambda n: "KREMES" in n),
    ("Ghee BB Booster",         lambda n: "GHEE" in n),
    ("Bubur 6+ Isi 3Cup @80ml", lambda n: "BUBUR" in n and "6+"  in n and "80" in n),
    ("Bubur 9+ Isi 3Cup @80ml", lambda n: "BUBUR" in n and "9+"  in n and "80" in n),
    ("Bubur 11+ Isi 3Cup @80ml",lambda n: "BUBUR" in n and "11+" in n and "80" in n),
]


def get_category(product_name: str) -> str:
    """Return label kategori jika produk cocok, atau None jika tidak ada yang cocok."""
    n = product_name.upper()
    for label, fn in CATEGORY_RULES:
        if fn(n):
            return label
    return None

# ─────────────────────────────────────────────
# 1. Baca data dari source file
# ─────────────────────────────────────────────
def load_data(filepath: str) -> list[dict]:
    """Baca semua baris produk dari source Excel, abaikan header & section rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    HEADER_ROW = 1
    SECTION_KEYWORDS = {"HERTASNING", "PERINTIS", "MALANG"}

    records = []
    current_outlet = ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku, nama, qty, harga_satuan, diskon, harga_jual, harga_rata, cogs, gp, gp_pct, pct_penjualan, metode = row[:12]

        # Deteksi baris section header
        if sku and qty is None and nama is None:
            if str(sku).strip().upper() in SECTION_KEYWORDS:
                current_outlet = str(sku).strip()
            continue

        # Skip baris kosong atau baris total (biasanya tidak ada SKU numerik)
        if not sku or not nama:
            continue

        # Konversi qty ke angka
        try:
            qty_val = float(str(qty).replace(",", "").strip())
        except (ValueError, TypeError):
            continue

        try:
            harga_jual_val = float(str(harga_jual).replace(",", "").strip()) if harga_jual else 0
        except (ValueError, TypeError):
            harga_jual_val = 0

        try:
            gp_val = float(str(gp).replace(",", "").strip()) if gp else 0
        except (ValueError, TypeError):
            gp_val = 0

        # Skip produk yang gross profit-nya <= 0 (termasuk yang rugi)
        if gp_val <= 0:
            continue

        try:
            gp_pct_str = str(gp_pct).replace("%", "").strip() if gp_pct else "0"
            gp_pct_val = float(gp_pct_str)
        except (ValueError, TypeError):
            gp_pct_val = 0

        nama_upper = str(nama).strip().upper()

        # Skip produk mainan
        if "MAINAN" in nama_upper:
            continue

        # Skip produk yang tidak masuk kategori yang diizinkan
        kategori = get_category(nama_upper)
        if kategori is None:
            continue

        records.append({
            "Outlet"                  : current_outlet,
            "SKU"                     : str(sku).strip(),
            "Produk"                  : str(nama).strip(),
            "Kategori"                : kategori,
            "Qty"                     : qty_val,
            "Harga Satuan"            : float(str(harga_satuan).replace(",", "").strip()) if harga_satuan else 0,
            "Diskon"                  : float(str(diskon).replace(",", "").strip()) if diskon else 0,
            "Harga Jual Setelah Diskon": harga_jual_val,
            "Gross Profit"            : gp_val,
            "Gross Profit %"          : gp_pct_val,
            "Metode Pembayaran"       : str(metode).strip() if metode else "",
        })

    return _aggregate_records(records)

def _aggregate_records(records: list[dict]) -> list[dict]:
    """
    Gabungkan penjualan berdasarkan Kategori.
    Lalu urutkan berdasarkan Qty terbanyak (Top 20).
    """
    from collections import defaultdict

    groups = defaultdict(list)
    for r in records:
        groups[r["Kategori"]].append(r)

    aggregated = []
    for cat, rows in groups.items():
        total_qty   = sum(r["Qty"] for r in rows)
        total_gp    = sum(r["Gross Profit"] for r in rows)
        total_hj    = sum(r["Harga Jual Setelah Diskon"] for r in rows)
        total_diskon = sum(r["Diskon"] for r in rows)
        gp_pct_agg  = (total_gp / total_hj * 100) if total_hj else 0

        aggregated.append({
            "Kategori"                : cat,
            "Qty"                     : total_qty,
            "Diskon"                  : total_diskon,
            "Harga Jual Setelah Diskon": total_hj,
            "Gross Profit"            : total_gp,
            "Gross Profit %"          : gp_pct_agg,
        })

    # Urutkan berdasarkan Qty Terjual terbanyak ke paling sedikit
    aggregated.sort(key=lambda x: x["Qty"], reverse=True)

    # Ambil Top 20
    return aggregated[:20]





# ─────────────────────────────────────────────
# 2. Style helpers
# ─────────────────────────────────────────────
def _side():
    return Side(style="thin", color="BFBFBF")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

# Number formats
FMT_INT   = '#,##0'
FMT_MONEY = '#,##0'
FMT_PCT   = '0.00"%"'


# ─────────────────────────────────────────────
# 3. Build Excel output
# ─────────────────────────────────────────────
def build_excel(records: list[dict], title: str, output_path: str, source_file: str = ""):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # hapus default sheet

    # ── Sheet 1: Semua Outlet Gabungan ───────
    ws = wb.create_sheet("Ringkasan Gabungan")
    _write_sheet_content(ws, records, title)

    # ── Sheet 2: Data Mentah (salinan sumber) ─
    if source_file and os.path.exists(source_file):
        ws_raw = wb.create_sheet("Data Mentah")
        _copy_source_sheet(source_file, ws_raw)

    wb.save(output_path)
    print(f"[OK]  Saved: {output_path}")


def _copy_source_sheet(source_file: str, ws_dest):
    """Salin seluruh isi sheet aktif dari file sumber ke ws_dest."""
    wb_src = openpyxl.load_workbook(source_file, data_only=True)
    ws_src = wb_src.active

    # Salin lebar baris
    for row_idx, row_dim in ws_src.row_dimensions.items():
        ws_dest.row_dimensions[row_idx].height = row_dim.height

    # Salin nilai sel + hitung panjang teks terpanjang per kolom
    col_max_len: dict[int, int] = {}
    for row in ws_src.iter_rows():
        for cell in row:
            ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
            cell_len = len(str(cell.value)) if cell.value is not None else 0
            if cell_len > col_max_len.get(cell.column, 0):
                col_max_len[cell.column] = cell_len

    # Set lebar kolom otomatis berdasarkan teks terpanjang (min 8, max 60)
    for col_idx, max_len in col_max_len.items():
        adjusted = min(max(max_len + 2, 8), 60)
        ws_dest.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Gaya header baris pertama
    accent = "1A56DB"
    header_font_color = "FFFFFF"
    for cell in ws_dest[1]:
        if cell.value is not None:
            cell.font = Font(name="Calibri", bold=True, size=10, color=header_font_color)
            cell.fill = PatternFill("solid", fgColor=accent)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _border()

    # Freeze baris header
    ws_dest.freeze_panes = "A2"

    wb_src.close()
    print(f"[OK]  Sheet 'Data Mentah' disalin dari sumber.")


def _write_sheet_content(ws, records: list[dict], title: str):
    accent       = "1A56DB"   # biru
    accent_light = "EBF5FF"
    accent_alt   = "DBEAFE"
    header_font_color = "FFFFFF"

    # ── Baris judul ─────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color=header_font_color)
    title_cell.fill  = _header_fill(accent)
    title_cell.alignment = _center()

    # ── Baris subtitle ───────────────────────
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Ringkasan Penjualan per Kategori — Semua Outlet"
    sub_cell.font  = Font(name="Calibri", bold=True, size=11, color="374151")
    sub_cell.fill  = _header_fill("F3F4F6")
    sub_cell.alignment = _center()

    # ── Header kolom ────────────────────────
    headers = [
        "No", "Kategori",
        "Qty Terjual",
        "Diskon (Rp)", "Harga Jual Setelah Diskon (Rp)",
        "Gross Profit (Rp)", "Gross Profit %",
    ]
    header_row = 3
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=header_font_color)
        cell.fill      = _header_fill(accent)
        cell.alignment = _center()
        cell.border    = _border()

    # ── Data rows ───────────────────────────
    for idx, rec in enumerate(records, start=1):
        row_num   = header_row + idx
        fill_color = accent_light if idx % 2 == 1 else accent_alt

        values = [
            idx,
            rec["Kategori"],
            rec["Qty"],
            rec["Diskon"],
            rec["Harga Jual Setelah Diskon"],
            rec["Gross Profit"],
            rec["Gross Profit %"],
        ]

        formats = [
            None, None,
            FMT_INT, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_PCT,
        ]
        aligns = [
            _center(), _center(),
            _right(), _right(), _right(), _right(), _right(),
        ]

        for col_idx, (val, fmt, aln) in enumerate(zip(values, formats, aligns), start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill   = PatternFill("solid", fgColor=fill_color)
            cell.border = _border()
            cell.font   = Font(name="Calibri", size=10)
            if fmt:
                cell.number_format = fmt
            if aln:
                cell.alignment = aln
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Total row ───────────────────────────
    total_row_num = header_row + len(records) + 1
    
    total_qty = sum(r["Qty"] for r in records)
    total_diskon = sum(r["Diskon"] for r in records)
    total_hj = sum(r["Harga Jual Setelah Diskon"] for r in records)
    total_gp = sum(r["Gross Profit"] for r in records)
    total_gp_pct = (total_gp / total_hj * 100) if total_hj else 0

    total_values = [
        "", "TOTAL KESELURUHAN",
        total_qty, total_diskon, total_hj, total_gp, total_gp_pct
    ]

    for col_idx, (val, fmt, aln) in enumerate(zip(total_values, formats, aligns), start=1):
        cell = ws.cell(row=total_row_num, column=col_idx, value=val)
        cell.fill   = PatternFill("solid", fgColor="E5E7EB")  # Gray-200
        cell.border = _border()
        cell.font   = Font(name="Calibri", size=11, bold=True)
        if fmt:
            cell.number_format = fmt
        if aln:
            cell.alignment = aln
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Kolom width ─────────────────────────
    col_widths = [5, 30, 15, 20, 30, 22, 14]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row heights
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[header_row].height = 36
    for i in range(1, len(records) + 2):  # +2 to include total row
        ws.row_dimensions[header_row + i].height = 18

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"


# ─────────────────────────────────────────────
# 4. Main script
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate Ringkasan Penjualan KUNUKU BABY FOOD per Kategori (21 Kategori)")
    parser.add_argument("input_file", nargs="?", help="Path file Excel sumber dari Luna POS", default=None)
    parser.add_argument("-o", "--output", help="Path file output (opsional)", default="")
    args = parser.parse_args()

    source_file = args.input_file

    if not source_file:
        print("="*60)
        print("  PROGRAM RINGKASAN PENJUALAN KUNUKU BABY FOOD")
        print("="*60)
        source_file = input("Silakan masukkan path file Excel (bisa drag & drop ke sini): ").strip()
        
        # Bersihkan tanda kutip jika user drag & drop file di terminal
        if source_file.startswith('"') and source_file.endswith('"'):
            source_file = source_file[1:-1]
        elif source_file.startswith("'") and source_file.endswith("'"):
            source_file = source_file[1:-1]

    if not source_file:
        print("[ERROR] Path file tidak boleh kosong!")
        sys.exit(1)

    if not os.path.exists(source_file):
        print(f"[ERROR] File tidak ditemukan: {source_file}")
        sys.exit(1)

    print(f"[READ]  Membaca: {source_file} ...")
    records = load_data(source_file)
    if not records:
        print("[WARN]  Tidak ada data yang valid untuk diproses.")
        sys.exit(0)

    print(f"    Total kategori terdeteksi: {len(records)}")

    # Ekstrak tanggal/periode dari nama file jika memungkinkan
    # Contoh: "Ringkasan Penjualan Produk 01 April - 30 April 2026.xlsx" -> "01 April - 30 April 2026"
    filename = os.path.basename(source_file)
    title_suffix = ""
    match = re.search(r'(\d{1,2}\s+[a-zA-Z]+\s+(?:-\s+\d{1,2}\s+[a-zA-Z]+\s+)?\d{4})', filename)
    if match:
        title_suffix = f" — {match.group(1)}"
    
    title = f"Ringkasan Penjualan per Kategori{title_suffix}"

    # Tentukan nama output file
    if args.output:
        output_file = args.output
    else:
        # Generate otomatis
        name, ext = os.path.splitext(filename)
        output_file = f"KATEGORI_{name}{ext}"

    build_excel(records, title=title, output_path=output_file, source_file=source_file)

if __name__ == "__main__":
    main()
